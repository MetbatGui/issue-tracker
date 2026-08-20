"""유상증자 Excel -> SQLite 마이그레이션 스크립트 테스트

실제 CapitalIncreaseExcelWriter로 생성한 Excel을 입력으로 사용해, 마이그레이션 후
row count parity(고유 접수번호 수 == DB row 수)가 지켜지는지 회귀 테스트로 고정합니다.
"""
import sys
from datetime import date
from pathlib import Path

import pytest

sys.path.insert(0, str(Path(__file__).resolve().parent.parent.parent / "scripts"))

from src.domain import CapitalIncreaseDecision
from src.domain.value_objects import StockInfo, FundingPurpose
from src.infrastructure.capital_increase_excel_writer import CapitalIncreaseExcelWriter
from migrate_capital_increase_to_sqlite import migrate


def _make_decision(rcept_no: str, year: int, company_name: str = "테스트회사") -> CapitalIncreaseDecision:
    return CapitalIncreaseDecision(
        source_filename=f"{company_name}_{rcept_no}.xml",
        company_name=company_name,
        new_shares=StockInfo(common=1000, preferred=0),
        par_value=500,
        total_shares_before=10000,
        issue_price=1200,
        funding=FundingPurpose(facility=100, operating=200, acquisition=0, other=10),
        method="일반공모",
        assign_per_share=0.1,
        board_resolution_date=date(year, 1, 1),
        disclosure_date=date(year, 1, 2),
        record_date=date(year, 1, 10),
        subscription_date=date(year, 1, 15),
        payment_date=date(year, 1, 20),
        report_name="주요사항보고서(유상증자결정)",
        is_correction=False,
        rcept_no=rcept_no,
        parent_rcp_no=None,
        original_disclosure_date=None,
    )


@pytest.fixture
def sample_excel(tmp_path):
    excel_path = tmp_path / "유상증자.xlsx"
    writer = CapitalIncreaseExcelWriter(output_path=str(excel_path))
    decisions = [
        _make_decision("20230101000001", 2023, "회사A"),
        _make_decision("20240101000002", 2024, "회사B"),
        _make_decision("20240102000003", 2024, "회사C"),
    ]
    writer.write(decisions)
    return excel_path


class TestManualFieldsThroughFullMigrationPath:
    """final_issue_price/listing_date는 migrate()가 Excel 셀에서 직접 읽어와야 하는 값.
    repository 단위테스트는 '한번 채워지면 덮어쓰지 않는다'만 확인했으므로, migrate() 자체가
    Excel에 이미 채워진 값을 실제로 추출해 옮기는지는 별도로 검증해야 함.
    """

    def test_manually_filled_excel_cells_are_migrated_into_db(self, sample_excel, tmp_path):
        import openpyxl

        wb = openpyxl.load_workbook(sample_excel)
        ws = wb["2024"]
        header_row = 2  # startrow=1(0-indexed) -> 엑셀 2번째 행이 헤더
        headers = {cell.value: cell.column for cell in ws[header_row]}
        data_row = header_row + 1

        ws.cell(row=data_row, column=headers["발행확정가액"], value="1,250")
        ws.cell(row=data_row, column=headers["신주상장일"], value="2024-02-01")
        edited_rcept_no = str(ws.cell(row=data_row, column=headers["접수번호"]).value)
        wb.save(sample_excel)

        db_path = str(tmp_path / "유상증자_manual.db")
        migrate(str(sample_excel), db_path)

        from src.infrastructure.capital_increase_sqlite_repository import CapitalIncreaseSqliteRepository
        repo = CapitalIncreaseSqliteRepository(db_path)
        row = repo._conn.execute(
            "SELECT final_issue_price, listing_date FROM capital_increase_decisions WHERE rcept_no = ?",
            (edited_rcept_no,),
        ).fetchone()

        assert row == ("1,250", "2024-02-01")

    def test_untouched_rows_have_null_manual_fields(self, sample_excel, tmp_path):
        db_path = str(tmp_path / "유상증자.db")
        migrate(str(sample_excel), db_path)

        from src.infrastructure.capital_increase_sqlite_repository import CapitalIncreaseSqliteRepository
        repo = CapitalIncreaseSqliteRepository(db_path)
        row = repo._conn.execute(
            "SELECT final_issue_price, listing_date FROM capital_increase_decisions WHERE rcept_no = ?",
            ("20230101000001",),
        ).fetchone()

        assert row == (None, None)


class TestMigration:
    def test_row_count_parity(self, sample_excel, tmp_path):
        db_path = str(tmp_path / "유상증자.db")

        result = migrate(str(sample_excel), db_path)

        assert result["excel_rows"] == 3
        assert result["db_rows"] == 3
        assert result["match"] is True

    def test_migrated_data_is_queryable_via_repository(self, sample_excel, tmp_path):
        from src.infrastructure.capital_increase_sqlite_repository import CapitalIncreaseSqliteRepository

        db_path = str(tmp_path / "유상증자.db")
        migrate(str(sample_excel), db_path)

        repo = CapitalIncreaseSqliteRepository(db_path)
        decisions = {d.rcept_no: d for d in repo.get_all()}

        assert len(decisions) == 3
        assert decisions["20240101000002"].company_name == "회사B"
        assert decisions["20240101000002"].disclosure_date == date(2024, 1, 2)
        assert decisions["20240101000002"].funding.facility == 100
